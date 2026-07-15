"""
데모용 샘플 데이터 생성 스크립트.
Excel 3개 + PDF 2개, 익명화된 랜덤 이름 사용.
파일명 규칙: [목적] [지출월]-[총액]만원.[ext]
"""

import os
import random

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "data")

LAST_NAMES = ["김", "이", "박", "최", "정", "강", "조", "윤", "장", "임",
              "한", "오", "서", "신", "권", "황", "안", "송", "류", "홍"]
FIRST_NAMES_M = ["민준", "서준", "예준", "도윤", "시우", "주원", "하준",
                 "지호", "준서", "현우", "성민", "재원", "동현", "민혁"]
FIRST_NAMES_F = ["서연", "서윤", "지우", "서현", "민서", "하은", "하윤",
                 "윤서", "채원", "수아", "지민", "예은", "나연", "지현"]
FIRST_NAMES = FIRST_NAMES_M + FIRST_NAMES_F

DEPARTMENTS = ["전기과", "기계과", "건축과", "화학과", "컴퓨터과",
               "전자과", "자동화과", "친환경자동차과", "섬유소재과"]

FONT_PATH = r"C:\Windows\Fonts\malgun.ttf"


# ---------------------------------------------------------------------------
# 유틸
# ---------------------------------------------------------------------------
def random_name(used: set) -> str:
    for _ in range(100):
        name = random.choice(LAST_NAMES) + random.choice(FIRST_NAMES)
        if name not in used:
            used.add(name)
            return name
    return random.choice(LAST_NAMES) + random.choice(FIRST_NAMES)


def random_birthdate() -> str:
    year = random.randint(5, 9)
    month = random.randint(1, 12)
    day = random.randint(1, 28)
    return f"{year:02d}{month:02d}{day:02d}"


def make_rows(n: int, grade_list: list, amounts: list[int], recipient: str) -> list[dict]:
    """amounts: 학생 수와 같은 길이의 금액 리스트 (차등 지급 가능)."""
    assert len(amounts) == n, f"amounts 길이({len(amounts)})가 n({n})과 다릅니다"
    shuffled = amounts.copy()
    random.shuffle(shuffled)
    used: set = set()
    rows = []
    for i in range(n):
        rows.append({
            "연번": i + 1,
            "학과": random.choice(DEPARTMENTS),
            "학년": random.choice(grade_list),
            "성명": random_name(used),
            "생년월일": random_birthdate(),
            "금액": shuffled[i],
            "지급처": recipient,
        })
    return sorted(rows, key=lambda r: (-r["금액"], r["학과"], r["학년"]))


# ---------------------------------------------------------------------------
# Excel 생성
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center")


def write_excel(filename: str, title: str, description: str, rows: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "장학금 명단"

    # 행 1: 제목
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = _CENTER
    ws.row_dimensions[1].height = 24

    # 행 2~3: 설명 텍스트
    ws.merge_cells("A2:G3")
    desc_cell = ws["A2"]
    desc_cell.value = description
    desc_cell.font = Font(size=10)
    desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30

    # 행 4: 헤더
    headers = ["연번", "학과", "학년", "성명", "생년월일", "금액", "지급처"]
    col_widths = [6, 16, 6, 8, 12, 10, 20]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[4].height = 18

    # 행 5~: 데이터
    for row_idx, row in enumerate(rows, start=5):
        for col_idx, key in enumerate(["연번", "학과", "학년", "성명", "생년월일", "금액", "지급처"], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            cell.alignment = _CENTER
            cell.border = _BORDER

    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"  [Excel] {filename}  ({len(rows)}명)")


# ---------------------------------------------------------------------------
# PDF 생성
# ---------------------------------------------------------------------------
def write_pdf(filename: str, title: str, description: str, rows: list[dict]):
    pdfmetrics.registerFont(TTFont("Malgun", FONT_PATH))

    path = os.path.join(OUTPUT_DIR, filename)
    doc = SimpleDocTemplate(path, pagesize=A4,
                            leftMargin=40, rightMargin=40,
                            topMargin=50, bottomMargin=40)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", fontName="Malgun", fontSize=14,
                                 alignment=1, spaceAfter=12, leading=20)
    desc_style = ParagraphStyle("Desc", fontName="Malgun", fontSize=10,
                                leading=16, spaceAfter=12,
                                leftIndent=4, rightIndent=4)
    cell_style = ParagraphStyle("Cell", fontName="Malgun", fontSize=9, leading=12)

    headers = ["연번", "학과", "학년", "성명", "생년월일", "금액", "지급처"]
    col_widths = [30, 90, 35, 50, 70, 60, 130]

    table_data = [[Paragraph(h, cell_style) for h in headers]]
    for row in rows:
        table_data.append([
            Paragraph(str(row["연번"]), cell_style),
            Paragraph(row["학과"], cell_style),
            Paragraph(str(row["학년"]), cell_style),
            Paragraph(row["성명"], cell_style),
            Paragraph(row["생년월일"], cell_style),
            Paragraph(f"{row['금액']:,}", cell_style),
            Paragraph(row["지급처"], cell_style),
        ])

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, -1), "Malgun"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2FF")]),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    doc.build([
        Paragraph(title, title_style),
        Paragraph(description, desc_style),
        Spacer(1, 6),
        tbl,
    ])
    print(f"  [PDF]   {filename}  ({len(rows)}명)")


# ---------------------------------------------------------------------------
# 파일 정의
# ---------------------------------------------------------------------------
FILES = [
    {
        "filename":    "신입생 동문장학금 3월-480만원.xlsx",
        "title":       "2024학년도 신입생 동문장학금 지급 명단",
        "description": (
            "본 명단은 2024년 3월 한빛공업고등학교 동문회에서 신입생을 대상으로 지급한 장학금 수혜자 명단입니다.\n"
            "지급 목적: 신입생의 학업 의욕 고취 및 동문회 장학 사업의 일환으로, 입학한 신입생 전원에게 균등 지급합니다.\n"
            "지급 기관: 한빛공업고등학교 동문회  |  지급 기준: 당해 신입학생 전원 균등 지급  |  1인당 지급액: 200,000원"
        ),
        "format":    "xlsx",
        "grades":    [1],
        "amounts":   [200000] * 24,
        "recipient": "한빛공업고등학교 동문회",
    },
    {
        "filename":    "성적우수 장학금 상반기 6월-320만원.pdf",
        "title":       "2024학년도 상반기 성적우수 장학금 지급 명단",
        "description": (
            "본 명단은 2024년 상반기 한빛공업고등학교에서 학업 성취가 우수한 학생들에게 지급한 장학금 수혜자 명단입니다.\n"
            "지급 목적: 우수한 학업 성취를 이룬 학생들을 포상하고 지속적인 학업 의욕을 고취하기 위해 지급합니다.\n"
            "지급 기관: 한빛공업고등학교  |  선발 기준: 학기 성적 상위 학생 차등 선발  |  지급액: 상위 250,000원 / 하위 150,000원"
        ),
        "format":    "pdf",
        "grades":    [1, 2, 3],
        "amounts":   [250000] * 8 + [150000] * 8,
        "recipient": "한빛공업고등학교",
    },
    {
        "filename":    "성적우수 장학금 하반기 12월-280만원.pdf",
        "title":       "2024학년도 하반기 성적우수 장학금 지급 명단",
        "description": (
            "본 명단은 2024년 하반기 한빛공업고등학교에서 학업 성취가 우수한 학생들에게 지급한 장학금 수혜자 명단입니다.\n"
            "지급 목적: 하반기 학업 우수자를 포상하여 성취 동기를 강화하고 학업 분위기 향상에 기여합니다.\n"
            "지급 기관: 한빛공업고등학교  |  선발 기준: 학기 성적 상위 학생 차등 선발  |  지급액: 상위 250,000원 / 하위 150,000원"
        ),
        "format":    "pdf",
        "grades":    [1, 2, 3],
        "amounts":   [250000] * 7 + [150000] * 7,
        "recipient": "한빛공업고등학교",
    },
    {
        "filename":    "체육특기생 지원금 9월-150만원.xlsx",
        "title":       "2024학년도 체육특기생 지원금 지급 명단",
        "description": (
            "본 명단은 2024년 9월 한빛공업고등학교 체육부에서 교내 체육 활동에 적극 참여하는 체육특기생에게 지급한 지원금 수혜자 명단입니다.\n"
            "지급 목적: 학교 체육 종목 대표 학생들의 훈련 및 활동 비용을 지원하고 건전한 학교 체육 문화를 장려합니다.\n"
            "지급 기관: 한빛공업고등학교 체육부  |  지급 기준: 교내 체육 종목 참가 학생 균등 지급  |  1인당 지급액: 150,000원"
        ),
        "format":    "xlsx",
        "grades":    [1, 2, 3],
        "amounts":   [150000] * 10,
        "recipient": "한빛공업고등학교 체육부",
    },
    {
        "filename":    "학년말 성적우수 장학금 12월-200만원.xlsx",
        "title":       "2024학년도 학년말 성적우수 장학금 지급 명단",
        "description": (
            "본 명단은 2024년 12월 학년말을 맞아 한빛공업고등학교에서 해당 학년 성적 우수자를 대상으로 지급한 장학금 수혜자 명단입니다.\n"
            "지급 목적: 학년 전체 성적 우수 학생을 단계별로 포상하여 다음 학년도 학업 의욕을 고취하고 우수 학생을 격려합니다.\n"
            "지급 기관: 한빛공업고등학교  |  대상: 2·3학년  |  지급액: 1등급 300,000원 / 2등급 200,000원 / 3등급 100,000원"
        ),
        "format":    "xlsx",
        "grades":    [2, 3],
        "amounts":   [300000] * 3 + [200000] * 4 + [100000] * 3,
        "recipient": "한빛공업고등학교",
    },
    # 업로드 기능 테스트용 문서
    {
        "filename":    "취업특기생 지원금 11월-240만원.xlsx",
        "title":       "2024학년도 취업특기생 지원금 지급 명단",
        "description": (
            "본 명단은 2024년 11월 한빛공업고등학교 취업지원실에서 현장실습 및 취업 활동에 참여하는 취업특기생에게 지급한 지원금 수혜자 명단입니다.\n"
            "지급 목적: 졸업 예정 학생들의 현장실습 및 취업 준비 활동을 장려하고 조기 취업을 지원합니다.\n"
            "지급 기관: 한빛공업고등학교 취업지원실  |  대상: 3학년 취업활동자  |  1인당 지급액: 200,000원"
        ),
        "format":    "xlsx",
        "grades":    [3],
        "amounts":   [200000] * 12,
        "recipient": "한빛공업고등학교 취업지원실",
    },
]


# ---------------------------------------------------------------------------
# 실행
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    random.seed(42)

    print(f"데모 데이터 생성 → {OUTPUT_DIR}\n")
    for f in FILES:
        rows = make_rows(len(f["amounts"]), f["grades"], f["amounts"], f["recipient"])
        if f["format"] == "xlsx":
            write_excel(f["filename"], f["title"], f["description"], rows)
        else:
            write_pdf(f["filename"], f["title"], f["description"], rows)

    print(f"\n완료: {len(FILES)}개 파일 생성")
